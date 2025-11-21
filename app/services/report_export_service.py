"""Report export service for exporting reports to Excel, PDF, and CSV."""
from io import BytesIO, StringIO
from typing import List, Dict, Any, Optional
from decimal import Decimal
from datetime import date, datetime
import csv

from app.services.report_service import ReportData
from app.services.pdf_service import PDFService


class ReportExportService:
    """Service for exporting reports to Excel, PDF, and CSV formats."""
    
    def __init__(self):
        self.pdf_service = PDFService()
    
    def export_to_excel(
        self,
        report_data: ReportData,
        include_charts: bool = False
    ) -> bytes:
        """
        Export report to Excel format.
        
        Args:
            report_data: ReportData object to export
            include_charts: Whether to include charts (not implemented yet)
            
        Returns:
            Excel file bytes
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_data.title[:31]  # Excel sheet name limit
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Title row
        ws.merge_cells(f'A1:{get_column_letter(len(report_data.data[0]) if report_data.data else 1)}1')
        title_cell = ws['A1']
        title_cell.value = report_data.title
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Period info
        if report_data.period_start and report_data.period_end:
            ws['A2'] = f"Period: {report_data.period_start} to {report_data.period_end}"
            ws['A2'].font = Font(size=10, italic=True)
        
        # Data headers
        if report_data.data:
            headers = list(report_data.data[0].keys())
            ws.append(headers)
            
            # Style headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Data rows
            for row_data in report_data.data:
                row = []
                for header in headers:
                    value = row_data.get(header)
                    # Format values
                    if isinstance(value, Decimal):
                        row.append(float(value))
                    elif isinstance(value, (date, datetime)):
                        row.append(value.isoformat())
                    else:
                        row.append(value)
                ws.append(row)
            
            # Auto-adjust column widths
            for col_idx, header in enumerate(headers, 1):
                max_length = len(str(header))
                column_letter = get_column_letter(col_idx)
                for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Summary section
        if report_data.summary:
            summary_row = len(report_data.data) + 5 if report_data.data else 5
            ws.cell(row=summary_row, column=1).value = "Summary"
            ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
            
            summary_row += 1
            for key, value in report_data.summary.items():
                if key not in ['top_products', 'top_customers', 'suppliers_summary']:  # Skip nested summaries
                    ws.cell(row=summary_row, column=1).value = str(key).replace('_', ' ').title()
                    ws.cell(row=summary_row, column=1).font = Font(bold=True)
                    if isinstance(value, (int, float, Decimal)):
                        ws.cell(row=summary_row, column=2).value = float(value)
                    else:
                        ws.cell(row=summary_row, column=2).value = str(value)
                    summary_row += 1
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
    
    def export_to_pdf(
        self,
        report_data: ReportData,
        include_charts: bool = False
    ) -> BytesIO:
        """
        Export report to PDF format.
        
        Args:
            report_data: ReportData object to export
            include_charts: Whether to include charts (not implemented yet)
            
        Returns:
            BytesIO object containing PDF data
        """
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_RIGHT, TA_LEFT, TA_CENTER
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20*mm, leftMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)
        
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        story.append(Paragraph(report_data.title, title_style))
        story.append(Spacer(1, 6*mm))
        
        # Period info
        if report_data.period_start and report_data.period_end:
            period_text = f"Period: {report_data.period_start} to {report_data.period_end}"
            story.append(Paragraph(period_text, styles['Normal']))
            story.append(Spacer(1, 6*mm))
        
        # Data table
        if report_data.data:
            # Prepare table data
            headers = list(report_data.data[0].keys())
            table_data = [headers]
            
            for row_data in report_data.data:
                row = []
                for header in headers:
                    value = row_data.get(header)
                    if isinstance(value, Decimal):
                        row.append(f"{value:,.2f}")
                    elif isinstance(value, (date, datetime)):
                        row.append(value.strftime('%Y-%m-%d'))
                    elif value is None:
                        row.append('')
                    else:
                        row.append(str(value))
                table_data.append(row)
            
            # Create table
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                # Header row
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                
                # Data rows
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
            ]))
            
            story.append(table)
            story.append(Spacer(1, 12*mm))
        
        # Summary section
        if report_data.summary:
            story.append(Paragraph("Summary", styles['Heading2']))
            story.append(Spacer(1, 6*mm))
            
            summary_data = []
            for key, value in report_data.summary.items():
                if key not in ['top_products', 'top_customers', 'suppliers_summary']:
                    if isinstance(value, (int, float, Decimal)):
                        summary_data.append([str(key).replace('_', ' ').title(), f"{float(value):,.2f}"])
                    else:
                        summary_data.append([str(key).replace('_', ' ').title(), str(value)])
            
            if summary_data:
                summary_table = Table(summary_data, colWidths=[80*mm, 80*mm])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ]))
                story.append(summary_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def export_to_csv(
        self,
        report_data: ReportData
    ) -> str:
        """
        Export report to CSV format.
        
        Args:
            report_data: ReportData object to export
            
        Returns:
            CSV string
        """
        output = StringIO()
        writer = csv.writer(output)
        
        # Title
        writer.writerow([report_data.title])
        if report_data.period_start and report_data.period_end:
            writer.writerow([f"Period: {report_data.period_start} to {report_data.period_end}"])
        writer.writerow([])  # Empty row
        
        # Headers
        if report_data.data:
            headers = list(report_data.data[0].keys())
            writer.writerow(headers)
            
            # Data rows
            for row_data in report_data.data:
                row = []
                for header in headers:
                    value = row_data.get(header)
                    if isinstance(value, Decimal):
                        row.append(str(value))
                    elif isinstance(value, (date, datetime)):
                        row.append(value.isoformat())
                    elif value is None:
                        row.append('')
                    else:
                        row.append(str(value))
                writer.writerow(row)
            
            writer.writerow([])  # Empty row
        
        # Summary
        if report_data.summary:
            writer.writerow(['Summary'])
            for key, value in report_data.summary.items():
                if key not in ['top_products', 'top_customers', 'suppliers_summary']:
                    if isinstance(value, (int, float, Decimal)):
                        writer.writerow([str(key).replace('_', ' ').title(), f"{float(value):,.2f}"])
                    else:
                        writer.writerow([str(key).replace('_', ' ').title(), str(value)])
        
        return output.getvalue()

