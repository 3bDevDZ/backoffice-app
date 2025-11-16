"""Import/Export service for products and customers."""
import csv
import io
from typing import List, Dict, Any, Optional
from decimal import Decimal
from flask_babel import gettext as _


class ImportExportService:
    """Service for importing and exporting data (products, customers, etc.)."""
    
    @staticmethod
    def export_products_to_csv(products: List[Dict[str, Any]]) -> str:
        """
        Export products to CSV format.
        
        Args:
            products: List of product dictionaries
            
        Returns:
            CSV string
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            _('Code'),
            _('Name'),
            _('Description'),
            _('Price'),
            _('Cost'),
            _('Unit of Measure'),
            _('Barcode'),
            _('Status'),
            _('Categories')
        ])
        
        # Data rows
        for product in products:
            categories = ', '.join([cat.get('name', '') for cat in product.get('categories', [])])
            writer.writerow([
                product.get('code', ''),
                product.get('name', ''),
                product.get('description', ''),
                str(product.get('price', '0')),
                str(product.get('cost', '') or ''),
                product.get('unit_of_measure', ''),
                product.get('barcode', ''),
                product.get('status', ''),
                categories
            ])
        
        return output.getvalue()
    
    @staticmethod
    def export_products_to_excel(products: List[Dict[str, Any]]) -> bytes:
        """
        Export products to Excel format.
        
        Args:
            products: List of product dictionaries
            
        Returns:
            Excel file bytes
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _('Products')
        
        # Header
        headers = [
            _('Code'),
            _('Name'),
            _('Description'),
            _('Price'),
            _('Cost'),
            _('Unit of Measure'),
            _('Barcode'),
            _('Status'),
            _('Categories')
        ]
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for product in products:
            categories = ', '.join([cat.get('name', '') for cat in product.get('categories', [])])
            ws.append([
                product.get('code', ''),
                product.get('name', ''),
                product.get('description', ''),
                product.get('price', 0),
                product.get('cost') or '',
                product.get('unit_of_measure', ''),
                product.get('barcode', ''),
                product.get('status', ''),
                categories
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def import_products_from_csv(csv_content: str, category_map: Dict[str, int]) -> Dict[str, Any]:
        """
        Import products from CSV format.
        
        Args:
            csv_content: CSV file content as string
            category_map: Dictionary mapping category names to category IDs
            
        Returns:
            Dictionary with 'success' list and 'errors' list
        """
        reader = csv.DictReader(io.StringIO(csv_content))
        success = []
        errors = []
        
        for row_num, row in enumerate(reader, start=2):  # Start at 2 (row 1 is header)
            try:
                # Validate required fields
                if not row.get('Code') or not row.get('Name'):
                    errors.append({
                        'row': row_num,
                        'code': row.get('Code', ''),
                        'error': _('Code and Name are required')
                    })
                    continue
                
                # Parse price
                try:
                    price = Decimal(str(row.get('Price', '0')).replace(',', '.'))
                except (ValueError, TypeError):
                    price = Decimal('0')
                
                # Parse cost (optional)
                cost = None
                if row.get('Cost'):
                    try:
                        cost = Decimal(str(row.get('Cost')).replace(',', '.'))
                    except (ValueError, TypeError):
                        pass
                
                # Map categories
                category_ids = []
                if row.get('Categories'):
                    category_names = [c.strip() for c in row.get('Categories').split(',')]
                    for cat_name in category_names:
                        if cat_name in category_map:
                            category_ids.append(category_map[cat_name])
                        else:
                            errors.append({
                                'row': row_num,
                                'code': row.get('Code', ''),
                                'error': _('Category not found: {}').format(cat_name)
                            })
                
                if not category_ids:
                    errors.append({
                        'row': row_num,
                        'code': row.get('Code', ''),
                        'error': _('At least one valid category is required')
                    })
                    continue
                
                product_data = {
                    'code': row.get('Code').strip(),
                    'name': row.get('Name').strip(),
                    'description': row.get('Description', '').strip() or None,
                    'price': price,
                    'cost': cost,
                    'unit_of_measure': row.get('Unit of Measure', '').strip() or None,
                    'barcode': row.get('Barcode', '').strip() or None,
                    'status': row.get('Status', 'active').strip() or 'active',
                    'category_ids': category_ids
                }
                
                success.append(product_data)
                
            except Exception as e:
                errors.append({
                    'row': row_num,
                    'code': row.get('Code', ''),
                    'error': str(e)
                })
        
        return {
            'success': success,
            'errors': errors
        }
    
    @staticmethod
    def import_products_from_excel(excel_file: bytes, category_map: Dict[str, int]) -> Dict[str, Any]:
        """
        Import products from Excel format.
        
        Args:
            excel_file: Excel file bytes
            category_map: Dictionary mapping category names to category IDs
            
        Returns:
            Dictionary with 'success' list and 'errors' list
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel import. Install it with: pip install openpyxl")
        
        wb = openpyxl.load_workbook(io.BytesIO(excel_file))
        ws = wb.active
        
        # Read header row
        headers = [cell.value for cell in ws[1]]
        header_map = {h: idx for idx, h in enumerate(headers) if h}
        
        success = []
        errors = []
        
        # Read data rows
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            try:
                # Convert row to dictionary
                row_dict = {}
                for header, idx in header_map.items():
                    cell = row[idx] if idx < len(row) else None
                    row_dict[header] = cell.value if cell else ''
                
                # Validate required fields
                code = str(row_dict.get(_('Code') or 'Code', '')).strip()
                name = str(row_dict.get(_('Name') or 'Name', '')).strip()
                
                if not code or not name:
                    errors.append({
                        'row': row_num,
                        'code': code,
                        'error': _('Code and Name are required')
                    })
                    continue
                
                # Parse price
                try:
                    price_str = str(row_dict.get(_('Price') or 'Price', '0')).replace(',', '.')
                    price = Decimal(price_str)
                except (ValueError, TypeError):
                    price = Decimal('0')
                
                # Parse cost (optional)
                cost = None
                cost_str = row_dict.get(_('Cost') or 'Cost', '')
                if cost_str:
                    try:
                        cost = Decimal(str(cost_str).replace(',', '.'))
                    except (ValueError, TypeError):
                        pass
                
                # Map categories
                category_ids = []
                categories_str = row_dict.get(_('Categories') or 'Categories', '')
                if categories_str:
                    category_names = [c.strip() for c in str(categories_str).split(',')]
                    for cat_name in category_names:
                        if cat_name in category_map:
                            category_ids.append(category_map[cat_name])
                        else:
                            errors.append({
                                'row': row_num,
                                'code': code,
                                'error': _('Category not found: {}').format(cat_name)
                            })
                
                if not category_ids:
                    errors.append({
                        'row': row_num,
                        'code': code,
                        'error': _('At least one valid category is required')
                    })
                    continue
                
                product_data = {
                    'code': code,
                    'name': name,
                    'description': str(row_dict.get(_('Description') or 'Description', '')).strip() or None,
                    'price': price,
                    'cost': cost,
                    'unit_of_measure': str(row_dict.get(_('Unit of Measure') or 'Unit of Measure', '')).strip() or None,
                    'barcode': str(row_dict.get(_('Barcode') or 'Barcode', '')).strip() or None,
                    'status': str(row_dict.get(_('Status') or 'Status', 'active')).strip() or 'active',
                    'category_ids': category_ids
                }
                
                success.append(product_data)
                
            except Exception as e:
                errors.append({
                    'row': row_num,
                    'code': '',
                    'error': str(e)
                })
        
        return {
            'success': success,
            'errors': errors
        }

